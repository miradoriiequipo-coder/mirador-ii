from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import Optional
from .. import models
from ..database import get_db
from ..auth import require_admin

router = APIRouter(prefix="/finances", tags=["finances"])


def _get_tid(t: Optional[int], db: Session) -> Optional[int]:
    if t:
        return t
    active = db.query(models.Tournament).filter(models.Tournament.is_active == True).first()
    return active.id if active else None


def _eligible_players(tid: int, db: Session):
    """Jugadores que reciben nuevas deudas: activos y NO lesionados/inactivos."""
    return db.query(models.Player).filter(
        models.Player.tournament_id == tid,
        models.Player.is_active == True,
        models.Player.status == "activo",
    ).all()


@router.get("/summary")
def get_summary(t: Optional[int] = Query(None), db: Session = Depends(get_db)):
    tid = _get_tid(t, db)
    # Incluir todos los jugadores del torneo (activos, lesionados, inactivos)
    q = db.query(models.Player)
    if tid:
        q = q.filter(models.Player.tournament_id == tid)
    players = q.order_by(models.Player.player_number).all()

    result = []
    for p in players:
        deudas = [d for d in p.deudas if d.tournament_id == tid] if tid else p.deudas
        pagos  = [pay for pay in p.payments if pay.tournament_id == tid] if tid else p.payments

        deuda_insc  = sum(d.monto for d in deudas if d.tipo == "inscripcion")
        deuda_arb   = sum(d.monto for d in deudas if d.tipo == "arbitraje")
        deuda_manual= sum(d.monto for d in deudas if d.tipo == "manual")
        deuda_total = deuda_insc + deuda_arb + deuda_manual

        pago_insc   = sum(pay.amount for pay in pagos if pay.payment_type == "inscripcion")
        pago_arb    = sum(pay.amount for pay in pagos if pay.payment_type == "arbitraje")
        pago_otros  = sum(pay.amount for pay in pagos if pay.payment_type in ("manual", "ajuste"))
        pago_total  = pago_insc + pago_arb + pago_otros

        result.append({
            "player_id": p.id,
            "player_name": p.full_name,
            "player_number": p.player_number,
            "status": p.status or "activo",
            "is_active": p.is_active,
            "deuda_inscripcion": deuda_insc,
            "deuda_arbitraje": deuda_arb,
            "deuda_manual": deuda_manual,
            "deuda_total": deuda_total,
            "pago_inscripcion": pago_insc,
            "pago_arbitraje": pago_arb,
            "pagado_total": pago_total,
            "saldo_pendiente": deuda_total - pago_total,
            "deudas": [{"id":d.id,"tipo":d.tipo,"fase":d.fase,"monto":d.monto,"concepto":d.concepto,"created_at":d.created_at.isoformat()} for d in deudas],
            "payments": [{"id":pay.id,"tipo":pay.payment_type,"fase":pay.phase,"monto":pay.amount,"notas":pay.notes,"created_at":pay.created_at.isoformat()} for pay in pagos],
        })
    return result


@router.get("/configs")
def get_configs(t: Optional[int] = Query(None), db: Session = Depends(get_db)):
    tid = _get_tid(t, db)
    q_insc = db.query(models.InscripcionConfig)
    q_arb  = db.query(models.ArbitrajePhase)
    if tid:
        q_insc = q_insc.filter(models.InscripcionConfig.tournament_id == tid)
        q_arb  = q_arb.filter(models.ArbitrajePhase.tournament_id == tid)
    return {
        "inscripciones": [{"id":c.id,"total_amount":c.total_amount,"num_players":c.num_players,"amount_per_player":c.amount_per_player,"total_matches":c.total_matches,"notes":c.notes,"created_at":c.created_at.isoformat()} for c in q_insc.all()],
        "arbitrajes":    [{"id":a.id,"fase":a.fase,"num_games":a.num_games,"price_per_game":a.price_per_game,"num_players":a.num_players,"total_phase":a.total_phase,"amount_per_player":a.amount_per_player,"notes":a.notes,"created_at":a.created_at.isoformat()} for a in q_arb.all()],
    }


@router.post("/inscripcion-config")
def create_inscripcion_config(body: dict, t: Optional[int] = Query(None), db: Session = Depends(get_db), current_user=Depends(require_admin)):
    tid = _get_tid(t, db)
    if not tid:
        raise HTTPException(status_code=400, detail="No hay torneo activo")
    total_amount = float(body.get("total_amount", 0))
    total_matches = int(body.get("total_matches", 0)) or None
    if total_amount <= 0:
        raise HTTPException(status_code=400, detail="El monto total debe ser mayor a 0")

    # Solo jugadores activos (no lesionados ni inactivos)
    players = _eligible_players(tid, db)
    if not players:
        raise HTTPException(status_code=400, detail="No hay jugadores activos en este torneo")

    num_players = len(players)
    amount_per_player = round(total_amount / num_players, 0)

    config = models.InscripcionConfig(
        tournament_id=tid,
        total_amount=total_amount,
        num_players=num_players,
        amount_per_player=amount_per_player,
        total_matches=total_matches,
        notes=body.get("notes", ""),
    )
    db.add(config)
    db.flush()

    concepto = f"Inscripción — ${int(total_amount):,} ÷ {num_players} jugadores"
    if total_matches:
        concepto += f" ({total_matches} partidos totales)"

    for player in players:
        db.add(models.Deuda(
            tournament_id=tid,
            player_id=player.id,
            tipo="inscripcion",
            monto=amount_per_player,
            concepto=concepto,
            config_id=config.id,
            config_tipo="inscripcion",
        ))
    db.commit()
    return {
        "message": f"Inscripción configurada. {num_players} jugadores activos. Cada uno debe ${int(amount_per_player):,}",
        "amount_per_player": amount_per_player,
        "num_players": num_players,
        "excluded_note": "Jugadores lesionados e inactivos no recibieron esta deuda."
    }


@router.delete("/inscripcion-config/{config_id}")
def delete_inscripcion_config(config_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db.query(models.Deuda).filter(models.Deuda.config_id == config_id, models.Deuda.config_tipo == "inscripcion").delete()
    config = db.query(models.InscripcionConfig).filter(models.InscripcionConfig.id == config_id).first()
    if config:
        db.delete(config)
    db.commit()
    return {"message": "Configuración eliminada"}


@router.post("/arbitraje-phase")
def create_arbitraje_phase(body: dict, t: Optional[int] = Query(None), db: Session = Depends(get_db), current_user=Depends(require_admin)):
    tid = _get_tid(t, db)
    if not tid:
        raise HTTPException(status_code=400, detail="No hay torneo activo")
    fase = body.get("fase", "").strip()
    num_games = int(body.get("num_games", 0))
    price_per_game = float(body.get("price_per_game", 0))
    if not fase or num_games <= 0 or price_per_game <= 0:
        raise HTTPException(status_code=400, detail="Fase, partidos y precio son obligatorios")

    # Solo jugadores activos — lesionados e inactivos NO reciben esta deuda
    players = _eligible_players(tid, db)
    if not players:
        raise HTTPException(status_code=400, detail="No hay jugadores activos en este torneo")

    num_players = len(players)
    total_phase = num_games * price_per_game
    amount_per_player = round(total_phase / num_players, 0)

    phase_config = models.ArbitrajePhase(
        tournament_id=tid,
        fase=fase,
        num_games=num_games,
        price_per_game=price_per_game,
        num_players=num_players,
        total_phase=total_phase,
        amount_per_player=amount_per_player,
        notes=body.get("notes", ""),
    )
    db.add(phase_config)
    db.flush()

    concepto = f"Arbitraje {fase} — {num_games} partidos × ${int(price_per_game):,} ÷ {num_players} jugadores activos"
    for player in players:
        db.add(models.Deuda(
            tournament_id=tid,
            player_id=player.id,
            tipo="arbitraje",
            fase=fase,
            monto=amount_per_player,
            concepto=concepto,
            config_id=phase_config.id,
            config_tipo="arbitraje",
        ))
    db.commit()

    excluded = db.query(models.Player).filter(
        models.Player.tournament_id == tid,
        models.Player.status.in_(["lesionado", "inactivo"]),
    ).count()

    return {
        "message": f"{fase}: {num_players} jugadores activos deben ${int(amount_per_player):,} c/u",
        "amount_per_player": amount_per_player,
        "excluded_players": excluded,
        "excluded_note": f"{excluded} jugador(es) lesionado(s)/inactivo(s) no recibieron esta deuda." if excluded else ""
    }


@router.delete("/arbitraje-phase/{phase_id}")
def delete_arbitraje_phase(phase_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db.query(models.Deuda).filter(models.Deuda.config_id == phase_id, models.Deuda.config_tipo == "arbitraje").delete()
    phase = db.query(models.ArbitrajePhase).filter(models.ArbitrajePhase.id == phase_id).first()
    if phase:
        db.delete(phase)
    db.commit()
    return {"message": "Fase eliminada"}


@router.post("/deuda-manual")
def add_deuda_manual(body: dict, t: Optional[int] = Query(None), db: Session = Depends(get_db), current_user=Depends(require_admin)):
    """Agregar deuda manual a un jugador específico (jugador nuevo, ajuste, etc.)."""
    tid = _get_tid(t, db)
    player_id = int(body.get("player_id", 0))
    monto = float(body.get("monto", 0))
    concepto = body.get("concepto", "").strip()
    tipo = body.get("tipo", "manual")

    if not player_id or monto <= 0:
        raise HTTPException(status_code=400, detail="Jugador y monto son obligatorios")

    player = db.query(models.Player).filter(models.Player.id == player_id).first()
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")

    db.add(models.Deuda(
        tournament_id=tid,
        player_id=player_id,
        tipo=tipo if tipo in ("inscripcion","arbitraje","manual") else "manual",
        fase=body.get("fase"),
        monto=monto,
        concepto=concepto or f"Deuda manual — {player.full_name}",
        config_id=None,
        config_tipo="manual",
    ))
    db.commit()
    return {"message": f"Deuda de ${int(monto):,} asignada a {player.full_name}"}


@router.delete("/deuda/{deuda_id}")
def delete_deuda(deuda_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    deuda = db.query(models.Deuda).filter(models.Deuda.id == deuda_id).first()
    if not deuda:
        raise HTTPException(status_code=404, detail="Deuda no encontrada")
    db.delete(deuda)
    db.commit()
    return {"message": "Deuda eliminada"}


@router.post("/payment")
def add_payment(body: dict, t: Optional[int] = Query(None), db: Session = Depends(get_db), current_user=Depends(require_admin)):
    tid = _get_tid(t, db)
    player_id = int(body.get("player_id", 0))
    payment_type = body.get("payment_type", "")
    amount = float(body.get("amount", 0))
    if not player_id or not payment_type or amount <= 0:
        raise HTTPException(status_code=400, detail="Jugador, tipo y monto son obligatorios")
    player = db.query(models.Player).filter(models.Player.id == player_id).first()
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")
    db.add(models.Payment(
        tournament_id=tid,
        player_id=player_id,
        payment_type=payment_type,
        phase=body.get("phase"),
        amount=amount,
        notes=body.get("notes", ""),
    ))
    db.commit()
    return {"message": f"Pago de ${int(amount):,} registrado para {player.full_name}"}


@router.delete("/payment/{payment_id}")
def delete_payment(payment_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    payment = db.query(models.Payment).filter(models.Payment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    db.delete(payment)
    db.commit()
    return {"message": "Pago eliminado"}


# ── HISTORIAL DE MOVIMIENTOS POR JUGADOR ─────────────────────────
@router.get("/player/{player_id}/history")
def player_history(
    player_id: int,
    t: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    player = db.query(models.Player).filter(models.Player.id == player_id).first()
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")

    tid = _get_tid(t, db)

    # Deudas (lo que debe)
    deudas = db.query(models.Deuda).filter(
        models.Deuda.player_id == player_id,
        models.Deuda.tournament_id == tid,
    ).order_by(models.Deuda.created_at).all()

    # Pagos y ajustes (lo que ha pagado)
    pagos = db.query(models.Payment).filter(
        models.Payment.player_id == player_id,
        models.Payment.tournament_id == tid,
    ).order_by(models.Payment.payment_date).all()

    # Construir línea de tiempo unificada
    movimientos = []

    for d in deudas:
        movimientos.append({
            "id":       d.id,
            "tipo":     "deuda",
            "concepto": d.concepto or d.tipo,
            "fase":     d.fase,
            "monto":    d.monto,
            "fecha":    d.created_at.strftime("%d/%m/%Y") if d.created_at else "—",
            "fuente":   "Excel" if (d.config_tipo or "").startswith("xls_") else "Manual",
            "color":    "red",
        })

    for p in pagos:
        es_ajuste   = p.payment_type == "ajuste"
        es_negativo = p.amount < 0
        movimientos.append({
            "id":       p.id,
            "tipo":     "ajuste" if es_ajuste else "pago",
            "concepto": p.notes or p.payment_type,
            "fase":     p.phase,
            "monto":    p.amount,
            "fecha":    (p.payment_date or p.created_at).strftime("%d/%m/%Y") if (p.payment_date or p.created_at) else "—",
            "fuente":   "Excel" if "(Excel)" in (p.notes or "") else "Manual",
            "color":    "orange" if es_negativo else "green",
        })

    # Ordenar por fecha
    movimientos.sort(key=lambda x: x["fecha"])

    total_deuda  = sum(d.monto for d in deudas)
    total_pagado = sum(p.amount for p in pagos if p.amount > 0)
    total_ajustes_neg = sum(abs(p.amount) for p in pagos if p.amount < 0)
    saldo_pendiente = total_deuda - total_pagado + total_ajustes_neg

    return {
        "jugador":    {"id": player.id, "nombre": player.full_name, "numero": player.player_number},
        "resumen":    {
            "total_deuda":     total_deuda,
            "total_pagado":    total_pagado,
            "saldo_pendiente": max(0, saldo_pendiente),
        },
        "movimientos": movimientos,
    }


# ── AJUSTE MANUAL (positivo = cargo adicional, negativo = abono/crédito) ──
@router.post("/adjustment")
def add_adjustment(
    body: dict,
    t: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    from datetime import datetime
    tid = _get_tid(t, db)
    player_id   = int(body.get("player_id", 0))
    monto       = float(body.get("monto", 0))
    concepto    = body.get("concepto", "").strip()
    tipo_ajuste = body.get("tipo", "pago")  # "pago" | "cargo" | "ajuste"
    fecha_str   = body.get("fecha", "")

    if not player_id or monto == 0 or not concepto:
        raise HTTPException(status_code=400, detail="Jugador, monto y concepto son obligatorios")

    player = db.query(models.Player).filter(models.Player.id == player_id).first()
    if not player:
        raise HTTPException(status_code=404, detail="Jugador no encontrado")

    # Fecha del ajuste (hoy si no se especifica)
    fecha = datetime.now()
    if fecha_str:
        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
        except:
            pass

    if tipo_ajuste == "cargo":
        # Cargo adicional → nueva deuda
        db.add(models.Deuda(
            tournament_id=tid,
            player_id=player_id,
            tipo="ajuste",
            monto=abs(monto),
            concepto=concepto,
            config_tipo="ajuste_manual",
        ))
        msg = f"Cargo de ${abs(monto):,.0f} registrado para {player.full_name}"
    else:
        # Pago o abono → payment positivo
        # Ajuste negativo → payment negativo (corrección/crédito)
        amount = abs(monto) if tipo_ajuste == "pago" else -abs(monto)
        db.add(models.Payment(
            tournament_id=tid,
            player_id=player_id,
            payment_type="ajuste",
            amount=amount,
            notes=concepto,
            payment_date=fecha,
        ))
        msg = f"{'Abono' if amount > 0 else 'Ajuste'} de ${abs(amount):,.0f} registrado para {player.full_name}"

    db.commit()
    return {"message": msg}


# ── EXPORTAR EXCEL ────────────────────────────────────────────────
@router.get("/export-excel")
def export_excel(t: Optional[int] = Query(None), db: Session = Depends(get_db)):
    """Genera Excel con la misma estructura del archivo de cuentas."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse

    tid = _get_tid(t, db)
    if not tid:
        raise HTTPException(status_code=400, detail="No hay torneo activo")

    # Datos
    players = db.query(models.Player).filter(
        models.Player.tournament_id == tid,
        models.Player.is_active == True,
    ).order_by(models.Player.player_number).all()

    insc_cfg = db.query(models.InscripcionConfig).filter(
        models.InscripcionConfig.tournament_id == tid
    ).first()

    arb_phases = db.query(models.ArbitrajePhase).filter(
        models.ArbitrajePhase.tournament_id == tid
    ).order_by(models.ArbitrajePhase.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuentas"

    # Colores
    navy = "0D2137"
    lime = "C1F100"
    light_gray = "F2F2F2"
    white = "FFFFFF"

    def cell_style(cell, bold=False, bg=None, fg="000000", align="center", border=True):
        cell.font = Font(bold=bold, color=fg, size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if border:
            thin = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila 1: Info general
    ws.merge_cells("A1:G1")
    ws["A1"] = f"MIRADOR II FC — Cuentas del torneo"
    cell_style(ws["A1"], bold=True, bg=navy, fg=white, align="left")

    if insc_cfg:
        ws["H1"] = "Inscripción"
        ws["I1"] = insc_cfg.total_amount
        cell_style(ws["H1"], bold=True)
        cell_style(ws["I1"])
        ws["J1"] = "Partidos"
        ws["K1"] = len(arb_phases)
        cell_style(ws["J1"], bold=True)
        cell_style(ws["K1"])
        if arb_phases:
            ws["L1"] = "Arbitrajes"
            ws["M1"] = arb_phases[0].price_per_game if arb_phases else 0
            cell_style(ws["L1"], bold=True)
            cell_style(ws["M1"])

    # Fila 2: Cabeceras
    n_arb = len(arb_phases)
    headers = ["#", "Nombre Jugadores", "Documento", "Salud", "Teléfono", "Dorsal",
               "Abono Insc.", "Debe Insc."]
    for i in range(1, n_arb + 1):
        headers.append(f"P{i}")
    headers += ["Abono Arb.", "Debe Arb.", "Total Abonado"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell_style(cell, bold=True, bg=navy, fg=white)

    # Filas de jugadores
    for row_idx, p in enumerate(players, 3):
        pagos = [pay for pay in p.payments if pay.tournament_id == tid]
        pago_insc = sum(pay.amount for pay in pagos if pay.payment_type == "inscripcion")
        pago_arb  = sum(pay.amount for pay in pagos if pay.payment_type in ("arbitraje", "manual", "ajuste"))
        total_abonado = pago_insc + pago_arb

        deuda_insc = insc_cfg.amount_per_player if insc_cfg else 0
        debe_insc  = max(0, deuda_insc - pago_insc)

        bg_row = light_gray if row_idx % 2 == 0 else white

        row_data = [
            row_idx - 2,
            p.full_name,
            p.id_number or "",
            p.health_info or "",
            p.phone or "",
            p.player_number or "",
            pago_insc,
            debe_insc,
        ]

        # Partidos de arbitraje
        for arb in arb_phases:
            row_data.append(arb.amount_per_player)

        # Arb abono y debe
        total_arb_deuda = sum(a.amount_per_player for a in arb_phases)
        debe_arb = max(0, total_arb_deuda - pago_arb)
        row_data += [pago_arb, debe_arb, total_abonado]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            is_money = col_idx >= 7
            cell_style(cell, bg=bg_row, align="right" if is_money else ("left" if col_idx == 2 else "center"))
            if is_money and isinstance(val, (int, float)):
                cell.number_format = '#,##0'

    # Fila de totales
    total_row = len(players) + 3
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.merge_cells(f"A{total_row}:F{total_row}")
    cell_style(ws[f"A{total_row}"], bold=True, bg=navy, fg=white, align="center")

    total_cols = len(headers)
    for col_idx in range(7, total_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell = ws.cell(row=total_row, column=col_idx,
            value=f"=SUM({col_letter}3:{col_letter}{total_row-1})")
        cell_style(cell, bold=True, bg=lime, fg="000000")
        cell.number_format = '#,##0'

    # Ancho de columnas
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 8
    for col_idx in range(7, total_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cuentas_mirador_ii.xlsx"}
    )