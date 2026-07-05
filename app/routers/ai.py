from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from typing import Optional
import os, base64, json, io
from groq import Groq
from .. import models
from ..database import get_db
from ..auth import require_admin

router = APIRouter(prefix="/ai", tags=["ai"])

GROQ_KEY   = os.getenv("GEMINI_API_KEY", "")
GROQ_MODEL = "llama-3.3-70b-versatile"


def _client():
    if not GROQ_KEY:
        raise HTTPException(status_code=500, detail="API KEY no configurada")
    return Groq(api_key=GROQ_KEY)


def _generate(prompt: str) -> str:
    c = _client()
    try:
        resp = c.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1024,
            temperature=0.7,
        )
        return resp.choices[0].message.content
    except Exception as e:
        err = str(e)
        if "rate_limit_exceeded" in err or "429" in err:
            # Extraer tiempo de espera si está disponible
            import re
            wait = re.search(r'try again in (\d+m[\d.]+s)', err)
            wait_msg = f" Intenta de nuevo en {wait.group(1)}." if wait else " Intenta en unos minutos."
            raise HTTPException(status_code=429, detail=f"⏳ Límite diario de IA alcanzado.{wait_msg} El límite se renueva cada 24 horas.")
        raise


def _generate_with_image(prompt: str, mime_type: str, image_data: bytes) -> str:
    c = _client()
    b64 = base64.b64encode(image_data).decode()
    resp = c.chat.completions.create(
        model="meta-llama/llama-4-scout-17b-16e-instruct",
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{b64}"}},
                {"type": "text", "text": prompt},
            ],
        }],
        max_tokens=1024,
    )
    return resp.choices[0].message.content


def _get_tournament_context(tid: Optional[int], db: Session) -> dict:
    from sqlalchemy.orm import joinedload
    from datetime import datetime
    if tid:
        t = db.query(models.Tournament).filter(models.Tournament.id == tid).first()
    else:
        t = db.query(models.Tournament).filter(models.Tournament.is_active == True).first()

    if not t:
        return {}

    players = db.query(models.Player).filter(
        models.Player.tournament_id == t.id,
        models.Player.is_active == True,
    ).all()

    matches = db.query(models.Match).options(
        joinedload(models.Match.goals).joinedload(models.Goal.player),
        joinedload(models.Match.goals).joinedload(models.Goal.assist_player),
    ).filter(
        models.Match.tournament_id == t.id,
    ).order_by(models.Match.match_date.desc()).all()

    finances = []
    for p in players:
        deudas      = [d for d in p.deudas if d.tournament_id == t.id]
        pagos       = [pay for pay in p.payments if pay.tournament_id == t.id]
        deuda_total = sum(d.monto for d in deudas)
        pagado      = sum(pay.amount for pay in pagos if pay.amount > 0)
        ajustes_neg = sum(abs(pay.amount) for pay in pagos if pay.amount < 0)
        saldo       = max(0, deuda_total - pagado + ajustes_neg)

        # Historial detallado de movimientos
        movimientos = []
        for d in sorted(deudas, key=lambda x: x.created_at or datetime.min):
            movimientos.append({
                "tipo":     "deuda",
                "concepto": d.concepto or d.tipo,
                "monto":    round(d.monto, 2),
                "fecha":    d.created_at.strftime("%d/%m/%Y") if d.created_at else "—",
            })
        for pay in sorted(pagos, key=lambda x: (x.payment_date or x.created_at or datetime.min)):
            movimientos.append({
                "tipo":     "pago" if pay.amount > 0 else "ajuste_negativo",
                "concepto": pay.notes or pay.payment_type,
                "monto":    round(pay.amount, 2),
                "fecha":    (pay.payment_date or pay.created_at).strftime("%d/%m/%Y") if (pay.payment_date or pay.created_at) else "—",
            })

        finances.append({
            "jugador":       p.full_name,
            "numero":        p.player_number,
            "deuda_total":   round(deuda_total, 2),
            "pagado":        round(pagado, 2),
            "saldo_pendiente": round(saldo, 2),
            "historial":     movimientos,
        })

    # Goleadores acumulados del torneo
    goleadores = {}
    asistentes = {}
    for m in matches:
        for g in m.goals:
            nombre = g.player.full_name if g.player else "?"
            goleadores[nombre] = goleadores.get(nombre, 0) + g.count
            if g.assist_player:
                an = g.assist_player.full_name
                asistentes[an] = asistentes.get(an, 0) + 1

    top_goleadores = sorted(goleadores.items(), key=lambda x: -x[1])[:5]
    top_asistentes = sorted(asistentes.items(), key=lambda x: -x[1])[:5]

    return {
        "torneo": t.name, "temporada": t.season,
        "jugadores": [{"nombre": p.full_name, "numero": p.player_number, "estado": p.status} for p in players],
        "partidos": [
            {
                "rival":     m.opponent,
                "fecha":     m.match_date.strftime("%d/%m/%Y %H:%M"),
                "fase":      m.phase,
                "jugado":    m.is_played,
                "resultado": f"Mirador II {m.home_score} - {m.away_score} {m.opponent}" if m.is_played else "Pendiente",
                "goles":     [
                    {
                        "jugador":    g.player.full_name if g.player else "?",
                        "cantidad":   g.count,
                        "asistencia": g.assist_player.full_name if g.assist_player else None,
                    }
                    for g in m.goals
                ],
            }
            for m in matches
        ],
        "finanzas": finances,
        "goleadores_torneo":  [{"jugador": n, "goles": c} for n, c in top_goleadores],
        "asistentes_torneo":  [{"jugador": n, "asistencias": c} for n, c in top_asistentes],
    }


# ── PARSER DIRECTO DEL EXCEL DE MIRADOR II ────────────────────────
def _to_num(val) -> float:
    """Convierte valores del Excel a float exacto en pesos."""
    if val is None or val == '-' or val == '':
        return 0.0
    if isinstance(val, str):
        if val.startswith('='):
            return 0.0
        val = val.replace(',', '.').strip()
        try:
            return float(val)
        except:
            return 0.0
    if isinstance(val, (float, int)):
        return float(val)
    return 0.0


def _limpiar_nombre(nombre: str) -> str:
    nombre = nombre.strip()
    partes = nombre.split(' ', 1)
    if partes[0].replace('.', '').isdigit() and len(partes) > 1:
        return partes[1].strip()
    return nombre


def parse_mirador_excel(content: bytes) -> dict:
    """
    Lee el Excel de Mirador II con data_only=True para obtener
    valores calculados en vez de fórmulas.

    Columnas (0-indexed):
      1  → # camiseta
      2  → nombre
      3  → inscripción total (fórmula calculada)
      4  → inscripción abono
      6  → arb F1 total (fórmula calculada)
      7  → arb F1 abono
      9  → arb F2 total (fórmula calculada)
      10 → arb F2 abono
      12 → arb F3 total (fórmula calculada)
      13 → arb F3 abono
      16 → TOTAL PAGADO (efectivo real recibido)
      17 → TOTAL X PERSONA (total que debe pagar)
      18 → TOTAL DEBE (lo que falta)
    """
    import openpyxl
    # data_only=True lee valores calculados, no fórmulas
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    jugadores  = []
    retirados  = []
    en_retirados = False

    for i, row in enumerate(rows):
        if i < 3:  # saltar filas 1-3 (título y encabezados)
            continue

        # Detectar sección de retirados
        texto_fila = " ".join(str(c) for c in row if c)
        if any(p in texto_fila.upper() for p in ["RETIRADO", "EXPULSADO", "LESIONADO"]):
            en_retirados = True
            continue

        # Saltar filas de configuración (tienen texto como "# JUGADORES", "ARBITRAJE", etc.)
        PALABRAS_CONFIG = ["JUGADORES", "ARBITRAJE", "VALOR", "CANTIDAD", "TOTAL ARBITRAJE",
                           "TOTAL TORNEO", "RECAUDO", "SALDO", "ABONO INSCRIPCION", "PAGOS"]
        if any(p in texto_fila.upper() for p in PALABRAS_CONFIG):
            continue

        nombre_raw = row[2] if len(row) > 2 else None
        if not nombre_raw or not isinstance(nombre_raw, str) or not nombre_raw.strip():
            continue

        nombre = _limpiar_nombre(nombre_raw)
        numero_raw = row[1] if len(row) > 1 else None
        numero = int(round(numero_raw)) if numero_raw and isinstance(numero_raw, (int, float)) else 0

        def col(idx):
            return row[idx] if len(row) > idx else None

        insc_total  = _to_num(col(3))
        insc_abono  = _to_num(col(4))
        arb1_total  = _to_num(col(6))
        arb1_abono  = _to_num(col(7))
        arb2_total  = _to_num(col(9))
        arb2_abono  = _to_num(col(10))
        arb3_total  = _to_num(col(12))
        arb3_abono  = _to_num(col(13))
        total_pagado = _to_num(col(16))
        total_x_persona = _to_num(col(17))

        # Calcular debe por concepto
        insc_debe  = max(0, insc_total  - insc_abono)
        arb1_debe  = max(0, arb1_total  - arb1_abono)
        arb2_debe  = max(0, arb2_total  - arb2_abono)
        arb3_debe  = max(0, arb3_total  - arb3_abono)
        total_debe = max(0, total_x_persona - total_pagado)

        jugador = {
            "numero":  numero,
            "nombre":  nombre,
            "estado":  "retirado" if en_retirados else "activo",
            "inscripcion": {
                "total": insc_total,
                "abono": insc_abono,
                "debe":  insc_debe,
            },
            "arb_f1": {
                "total": arb1_total,
                "abono": arb1_abono,
                "debe":  arb1_debe,
            },
            "arb_f2": {
                "total": arb2_total,
                "abono": arb2_abono,
                "debe":  arb2_debe,
            },
            "arb_f3": {
                "total": arb3_total,
                "abono": arb3_abono,
                "debe":  arb3_debe,
            },
            "total_abono":   total_pagado,
            "total_debe":    total_debe,
            "total_persona": total_x_persona,
        }

        if en_retirados:
            retirados.append(jugador)
        else:
            jugadores.append(jugador)

    return {"jugadores": jugadores, "retirados": retirados}


# ── 1. ASISTENTE CHAT ─────────────────────────────────────────────
@router.post("/chat")
def chat(body: dict, t: Optional[int] = Query(None), db: Session = Depends(get_db)):
    question = body.get("question", "").strip()
    if not question:
        raise HTTPException(status_code=400, detail="Pregunta vacía")

    context = _get_tournament_context(t, db)

    # ── Detectar jugador específico y mostrar tarjeta ──
    finance_card = None
    palabras_finanzas = ["deuda","abono","pago","debe","dinero","cuenta","pagado","pendiente","saldo","finanza","inscripción","inscripcion","arbitraje","finanzas"]

    # Buscar siempre si hay un jugador mencionado con alta confianza
    try:
        active = db.query(models.Tournament).filter(
            models.Tournament.id == t if t else models.Tournament.is_active == True
        ).first()
        if active:
            tid = active.id
            players = db.query(models.Player).filter(
                models.Player.tournament_id == tid
            ).all()

            import unicodedata
            def norm(s):
                return ''.join(c for c in unicodedata.normalize('NFD', s.lower()) if unicodedata.category(c) != 'Mn')

            q_norm = norm(question)
            palabras_q = [w for w in q_norm.split() if len(w) >= 3 and w.isalpha()]

            jugador = None
            mejor_score = 0
            for p in players:
                nombre_norm = norm(p.full_name)
                partes_nombre = nombre_norm.split()
                score = sum(1 for parte in partes_nombre if parte in palabras_q)
                if score > mejor_score:
                    mejor_score = score
                    jugador = p

            # Mostrar tarjeta si:
            # - Score alto (≥2 partes del nombre coinciden), O
            # - Score=1 Y la pregunta tiene palabras de finanzas
            es_finanzas = any(p in question.lower() for p in palabras_finanzas)
            if jugador and (mejor_score >= 2 or (mejor_score >= 1 and es_finanzas)):
                deudas = db.query(models.Deuda).filter(
                    models.Deuda.player_id == jugador.id,
                    models.Deuda.tournament_id == tid
                ).all()
                pagos = db.query(models.Payment).filter(
                    models.Payment.player_id == jugador.id,
                    models.Payment.tournament_id == tid
                ).all()

                deuda_insc = sum(d.monto for d in deudas if d.tipo == "inscripcion")
                deuda_arb  = sum(d.monto for d in deudas if d.tipo == "arbitraje")
                pago_insc  = sum(p.amount for p in pagos if p.payment_type == "inscripcion")
                pago_arb   = sum(p.amount for p in pagos if p.payment_type == "arbitraje")
                pago_otros = sum(p.amount for p in pagos if p.payment_type in ("manual","ajuste"))
                total_pago = pago_insc + pago_arb + pago_otros
                total_deuda = deuda_insc + deuda_arb

                historial = sorted([{
                    "fecha": (p.payment_date or p.created_at).strftime("%d/%m/%Y") if (p.payment_date or p.created_at) else "—",
                    "concepto": p.notes or p.payment_type,
                    "monto": p.amount
                } for p in pagos if p.amount > 0], key=lambda x: x["fecha"], reverse=True)[:8]

                finance_card = {
                    "player_name": jugador.full_name,
                    "player_number": jugador.player_number,
                    "deuda_inscripcion": deuda_insc,
                    "deuda_arbitraje": deuda_arb,
                    "deuda_total": total_deuda,
                    "pago_inscripcion": pago_insc,
                    "pago_arbitraje": pago_arb,
                    "pago_total": total_pago,
                    "saldo_pendiente": total_deuda - total_pago,
                    "historial": historial,
                }
    except Exception:
        finance_card = None

    instruccion_ia = ""
    if finance_card:
        instruccion_ia = """IMPORTANTE: Los datos financieros del jugador se mostrarán en una TARJETA VISUAL separada.
Tu respuesta debe ser MUY CORTA: solo una o dos frases amigables presentando al jugador.
NO repitas los números ni las deudas porque ya aparecen en la tarjeta."""

    prompt = f"""Eres el asistente oficial del equipo de fútbol Mirador II FC de Colombia.
Responde en español colombiano, de forma amigable y directa.
Solo responde sobre el equipo, torneo, jugadores, partidos y finanzas.
{instruccion_ia}

DATOS ACTUALES DEL TORNEO:
{json.dumps(context, ensure_ascii=False, indent=2)}

PREGUNTA: {question}

{"Responde en máximo 1 frase corta y amigable. Los datos se muestran en la tarjeta." if finance_card else "Responde de forma conversacional. Máximo 2 párrafos cortos."}"""

    try:
        answer = _generate(prompt)
        return {"answer": answer, "finance_card": finance_card}
    except Exception as e:
        msg = str(e)
        if "rate_limit_exceeded" in msg or "429" in msg:
            import re
            wait = re.search(r'try again in (\d+m[\d.]+s)', msg)
            wt = f" Espera {wait.group(1)}." if wait else " Intenta en unos minutos."
            raise HTTPException(status_code=429, detail=f"⏳ Límite diario de IA alcanzado.{wt}")
        raise HTTPException(status_code=500, detail=f"Error de IA: {msg}")


# ── 2. CRÓNICA DEL PARTIDO ────────────────────────────────────────
@router.post("/chronicle/{match_id}")
def generate_chronicle(match_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    from sqlalchemy.orm import joinedload
    match = db.query(models.Match).options(
        joinedload(models.Match.goals).joinedload(models.Goal.player),
        joinedload(models.Match.goals).joinedload(models.Goal.assist_player),
    ).filter(models.Match.id == match_id).first()

    if not match:
        raise HTTPException(status_code=404, detail="Partido no encontrado")
    if not match.is_played:
        raise HTTPException(status_code=400, detail="El partido aún no se ha jugado")

    resultado = "empate"
    if match.home_score > match.away_score:
        resultado = "victoria"
    elif match.home_score < match.away_score:
        resultado = "derrota"

    goles_info = []
    for g in match.goals:
        info = f"{g.player.full_name} ({g.count} gol{'es' if g.count > 1 else ''})"
        if g.assist_player:
            info += f" con asistencia de {g.assist_player.full_name}"
        goles_info.append(info)

    prompt = f"""Eres el cronista oficial del equipo Mirador II FC de fútbol amateur colombiano.
Escribe una crónica emocionante y divertida para publicar en Instagram.

DATOS DEL PARTIDO:
- Rival: {match.opponent}
- Resultado: Mirador II {match.home_score} - {match.away_score} {match.opponent}
- Desenlace: {resultado.upper()}
- Fase: {match.phase or 'Partido'}
- Lugar: {match.location or 'Por definir'}
- Fecha: {match.match_date.strftime('%d/%m/%Y')}
- Goleadores: {', '.join(goles_info) if goles_info else 'Sin goles registrados'}

INSTRUCCIONES:
- Español colombiano, informal y apasionado
- Máximo 5 líneas, perfecto para Instagram
- Incluye 3-5 emojis relevantes
- Termina con: #MiradorII #FutbolAficionado #Colombia
- Victoria: celebra con energía. Derrota: anima con positivismo. Empate: resalta el esfuerzo.
- NO uses comillas ni asteriscos"""

    try:
        chronicle = _generate(prompt)
        return {
            "chronicle": chronicle,
            "match": {
                "opponent": match.opponent,
                "score": f"Mirador II {match.home_score} - {match.away_score} {match.opponent}",
                "phase": match.phase,
                "date": match.match_date.strftime("%d/%m/%Y"),
            }
        }
    except Exception as e:
        msg = str(e)
        if "rate_limit_exceeded" in msg or "429" in msg:
            import re
            wait = re.search(r'try again in (\d+m[\d.]+s)', msg)
            wt = f" Espera {wait.group(1)}." if wait else " Intenta en unos minutos."
            raise HTTPException(status_code=429, detail=f"⏳ Límite diario de IA alcanzado.{wt}")
        raise HTTPException(status_code=500, detail=f"Error de IA: {msg}")


# ── 3. LECTOR DE ARCHIVO (Excel directo o Imagen con IA) ─────────
@router.post("/read-file")
async def read_file(
    file: UploadFile = File(...),
    t: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    content  = await file.read()
    filename = file.filename.lower()

    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            # ── Leer config del torneo ──
            def cel(row, idx):
                v = row[idx] if len(row) > idx else None
                if isinstance(v, float): return int(round(v))
                return v

            def num(row, idx):
                v = cel(row, idx)
                try: return float(v) if v else 0.0
                except: return 0.0

            # Obtener jugadores ya en la app para marcar nuevos
            if t:
                tid_check = t
            else:
                active = db.query(models.Tournament).filter(models.Tournament.is_active == True).first()
                tid_check = active.id if active else None

            existentes_cedulas = set()
            existentes_numeros = set()
            if tid_check:
                players_existentes = db.query(models.Player).filter(
                    models.Player.tournament_id == tid_check
                ).all()
                existentes_cedulas = {p.id_number for p in players_existentes if p.id_number}
                existentes_numeros = {p.player_number for p in players_existentes if p.player_number}

            # ── Columnas (0-indexed):
            # 0=#, 1=Nombre, 2=vacío, 3=Documento, 4=Salud, 5=Teléfono, 6=Dorsal
            # 7=Abono Insc, 8=Debe Insc
            # 9=P1, 10=P2, 11=P3, 12=P4, 13=P5, 14=P6
            # 15=Abono Arb, 16=Debe Arb, 17=Total

            jugadores = []
            tipo_actual = "antiguo"  # cambia a "nuevo" cuando encuentra fila "Nuevos"

            for i, row in enumerate(rows):
                if i < 7: continue  # saltar cabeceras (filas 1-7)
                if not row or not any(v for v in row): continue

                # Detectar sección Nuevos/Antiguos
                primer_cel = str(row[0] or '').strip().lower()
                if 'antiguo' in primer_cel:
                    tipo_actual = 'antiguo'; continue
                if 'nuevo' in primer_cel:
                    tipo_actual = 'nuevo'; continue

                # Fila de jugador: col 1 tiene nombre
                nombre_raw = row[1] if len(row) > 1 else None
                if not nombre_raw or not isinstance(nombre_raw, str) or not nombre_raw.strip():
                    continue

                nombre  = nombre_raw.strip()
                cedula  = str(cel(row, 3) or '').strip()
                salud   = str(cel(row, 4) or '').strip()
                telefono = str(cel(row, 5) or '').strip()
                dorsal  = cel(row, 6)
                try: dorsal = int(dorsal) if dorsal else 0
                except: dorsal = 0

                # Abonos e inscripción
                abono_insc = num(row, 7)
                debe_insc  = num(row, 8)
                total_insc = abono_insc + debe_insc

                # Partidos de arbitraje (columnas 9-14) — monto por partido
                partidos = []
                for p_idx in range(9, 15):
                    monto_partido = num(row, p_idx)
                    if monto_partido > 0:
                        partidos.append(int(round(monto_partido)))

                abono_arb = int(round(num(row, 15)))
                debe_arb  = int(round(num(row, 16)))
                total_arb = abono_arb + debe_arb
                total_abonado = abono_insc + abono_arb

                # Determinar si es nuevo en la app
                es_nuevo = cedula not in existentes_cedulas and dorsal not in existentes_numeros

                # Construir lista de abonos
                abonos = []
                if total_insc > 0 or debe_insc > 0:
                    abonos.append({
                        "concepto": "Inscripción",
                        "tipo": "inscripcion",
                        "fase": None,
                        "monto_total": int(round(total_insc)),
                        "abono": int(round(abono_insc)),
                    })
                # Un abono por partido (deuda individual) + abono total sin dividir
                for p_num, monto in enumerate(partidos, 1):
                    abonos.append({
                        "concepto": f"Arbitraje Partido {p_num}",
                        "tipo": "arbitraje",
                        "fase": f"Partido {p_num}",
                        "monto_total": monto,
                        "abono": 0,  # El abono total se registra aparte
                    })
                # Registrar el abono de arbitraje como un solo pago total (evita errores de redondeo)
                if abono_arb > 0:
                    abonos.append({
                        "concepto": "Abono Arbitrajes",
                        "tipo": "arbitraje",
                        "fase": None,
                        "monto_total": 0,  # Solo es pago, no deuda adicional
                        "abono": abono_arb,
                    })

                jugadores.append({
                    "numero": dorsal,
                    "nombre": nombre,
                    "cedula": cedula,
                    "salud": salud,
                    "telefono": telefono,
                    "tipo": tipo_actual,
                    "es_nuevo_en_app": es_nuevo,
                    "abonos": abonos,
                    "total_abonado": round(total_abonado),
                    "notas": "",
                })

            resumen = {
                "total_jugadores": len(jugadores),
                "jugadores_nuevos_en_app": sum(1 for j in jugadores if j["es_nuevo_en_app"]),
                "jugadores_actualizados": sum(1 for j in jugadores if not j["es_nuevo_en_app"]),
                "antiguos": sum(1 for j in jugadores if j["tipo"] == "antiguo"),
                "nuevos_equipo": sum(1 for j in jugadores if j["tipo"] == "nuevo"),
            }

            return {
                "success": True,
                "formato": "excel_mirador",
                "jugadores": jugadores,
                "retirados": [],
                "total": len(jugadores),
                "resumen": resumen,
            }

        else:
            raise HTTPException(status_code=400, detail="Solo se aceptan Excel (.xlsx) o imágenes")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")


# ── 4. IMPORTAR FINANZAS COMPLETAS (SYNC INTELIGENTE) ─────────────
@router.post("/import-finances")
def import_finances(
    body: dict,
    t: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    jugadores_data  = body.get("jugadores", [])
    retirados_data  = body.get("retirados", [])
    incluir_ret     = body.get("incluir_retirados", False)
    limpiar_primero = body.get("limpiar_primero", False)  # nuevo: borrar todo antes

    todos = jugadores_data + (retirados_data if incluir_ret else [])
    if not todos:
        raise HTTPException(status_code=400, detail="No hay jugadores para importar")

    if t:
        tid = t
    else:
        active = db.query(models.Tournament).filter(models.Tournament.is_active == True).first()
        if not active:
            raise HTTPException(status_code=400, detail="No hay torneo activo")
        tid = active.id

    # ── Opción limpiar: borrar TODOS los jugadores del torneo creados por Excel ──
    if limpiar_primero:
        players_existentes = db.query(models.Player).filter(
            models.Player.tournament_id == tid
        ).all()
        for p in players_existentes:
            db.query(models.Deuda).filter(models.Deuda.player_id == p.id).delete()
            db.query(models.Payment).filter(models.Payment.player_id == p.id).delete()
            db.delete(p)
        db.flush()

    creados = 0; actualizados = 0; errores = []

    for j in todos:
        try:
            nombre   = str(j.get("nombre", "")).strip()
            numero   = int(j.get("numero", 0)) if j.get("numero") else 0
            cedula   = str(j.get("cedula", "") or "").strip()
            telefono = str(j.get("telefono", "") or "").strip()
            salud    = str(j.get("salud", "") or "").strip()
            if not nombre:
                continue

            # ── Buscar SOLO por cédula o dorsal (no por nombre — muy impreciso) ──
            player = None
            if cedula and cedula not in ("0", ""):
                player = db.query(models.Player).filter(
                    models.Player.tournament_id == tid,
                    models.Player.id_number == cedula,
                ).first()
            if not player and numero:
                player = db.query(models.Player).filter(
                    models.Player.tournament_id == tid,
                    models.Player.player_number == numero,
                ).first()

            if player:
                player.full_name  = nombre
                if numero:   player.player_number = numero
                if cedula:   player.id_number = cedula
                if telefono: player.phone = telefono
                if salud:    player.health_info = salud
                player.is_active = True
                player.status    = "activo"
                actualizados += 1
            else:
                player = models.Player(
                    tournament_id=tid,
                    full_name=nombre,
                    player_number=numero or 0,
                    id_number=cedula or None,
                    phone=telefono or None,
                    health_info=salud or None,
                    is_active=True,
                    status="activo",
                )
                db.add(player)
                db.flush()
                creados += 1

            # ── Borrar TODO lo importado de Excel anteriormente ──
            db.query(models.Deuda).filter(
                models.Deuda.player_id == player.id,
                models.Deuda.config_tipo.like("xls_%"),
            ).delete(synchronize_session=False)
            db.query(models.Payment).filter(
                models.Payment.player_id == player.id,
                models.Payment.tournament_id == tid,
                models.Payment.notes.like("%(Excel)%"),
            ).delete(synchronize_session=False)
            db.flush()

            # ── Insertar deudas y pagos EXACTOS del Excel ──
            abonos = j.get("abonos", [])

            for idx, ab in enumerate(abonos):
                concepto    = ab.get("concepto", f"Concepto {idx+1}")
                tipo_pago   = ab.get("tipo", "manual")
                fase        = ab.get("fase", None)
                monto_total = int(round(float(ab.get("monto_total", 0) or 0)))
                abono_val   = int(round(float(ab.get("abono", 0) or 0)))

                # Registrar deuda exacta
                if monto_total > 0:
                    db.add(models.Deuda(
                        tournament_id=tid, player_id=player.id,
                        tipo=tipo_pago if tipo_pago in ("inscripcion","arbitraje") else "manual",
                        fase=fase, monto=monto_total, concepto=concepto,
                        config_tipo=f"xls_ab_{idx}",
                    ))

                # Registrar abono exacto (solo si mayor a 0)
                if abono_val > 0:
                    db.add(models.Payment(
                        tournament_id=tid, player_id=player.id,
                        payment_type=tipo_pago if tipo_pago in ("inscripcion","arbitraje") else "manual",
                        phase=fase, amount=abono_val,
                        notes=f"{concepto} (Excel)",
                    ))

        except Exception as e:
            errores.append(f"{j.get('nombre','?')}: {str(e)}")

    db.commit()
    return {
        "message": f"✅ {creados} jugadores nuevos · {actualizados} actualizados · {len(errores)} errores",
        "creados": creados, "actualizados": actualizados, "errores": errores,
    }
def import_finances(
    body: dict,
    t: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    jugadores_data  = body.get("jugadores", [])
    retirados_data  = body.get("retirados", [])
    incluir_ret     = body.get("incluir_retirados", False)

    todos = jugadores_data + (retirados_data if incluir_ret else [])
    if not todos:
        raise HTTPException(status_code=400, detail="No hay jugadores para importar")

    if t:
        tid = t
    else:
        active = db.query(models.Tournament).filter(models.Tournament.is_active == True).first()
        if not active:
            raise HTTPException(status_code=400, detail="No hay torneo activo")
        tid = active.id

    creados = 0; actualizados = 0; errores = []

    for j in todos:
        try:
            nombre = str(j.get("nombre", "")).strip()
            numero = int(j.get("numero", 0)) if j.get("numero") else 0
            cedula = str(j.get("cedula", "") or "").strip()
            telefono = str(j.get("telefono", "") or "").strip()
            tipo   = j.get("tipo", "activo")  # antiguo / nuevo
            if not nombre:
                continue

            # ── Buscar jugador existente ──
            player = None
            # 1. Por cédula
            if cedula and cedula not in ("0", ""):
                player = db.query(models.Player).filter(
                    models.Player.tournament_id == tid,
                    models.Player.id_number == cedula,
                ).first()
            # 2. Por número de camiseta
            if not player and numero:
                player = db.query(models.Player).filter(
                    models.Player.tournament_id == tid,
                    models.Player.player_number == numero,
                ).first()
            # 3. Por nombre aproximado
            if not player:
                palabras = nombre.split()
                for palabra in palabras:
                    if len(palabra) > 3:
                        player = db.query(models.Player).filter(
                            models.Player.tournament_id == tid,
                            models.Player.full_name.ilike(f"%{palabra}%"),
                        ).first()
                        if player:
                            break

            if player:
                # Actualizar datos y asegurar que esté activo
                player.full_name = nombre
                if numero: player.player_number = numero
                if cedula: player.id_number = cedula
                if telefono: player.phone = telefono
                player.is_active = True
                player.status = "activo"
                if j.get("salud"): player.health_info = j.get("salud")
                actualizados += 1
            else:
                # Crear jugador nuevo
                player = models.Player(
                    tournament_id=tid,
                    full_name=nombre,
                    player_number=numero or 0,
                    id_number=cedula or f"XLS-{tid}-{nombre[:8].replace(' ','')}",
                    phone=telefono or None,
                    health_info=j.get("salud") or None,
                    is_active=True,
                    status="activo",
                )
                db.add(player)
                db.flush()
                creados += 1

            # ── SYNC: borrar pagos y deudas anteriores importados del Excel ──
            db.query(models.Deuda).filter(
                models.Deuda.player_id == player.id,
                models.Deuda.config_tipo.like("xls_%"),
            ).delete(synchronize_session=False)

            db.query(models.Payment).filter(
                models.Payment.player_id == player.id,
                models.Payment.notes.like("%(Excel)%"),
            ).delete(synchronize_session=False)

            db.flush()

            # ── Insertar abonos del nuevo formato (lista de abonos) ──
            abonos = j.get("abonos", [])

            if abonos:
                # Nuevo formato flexible con lista de abonos
                for idx, ab in enumerate(abonos):
                    concepto   = ab.get("concepto", f"Concepto {idx+1}")
                    tipo_pago  = ab.get("tipo", "manual")
                    fase       = ab.get("fase", None)
                    monto_total = float(ab.get("monto_total", 0) or 0)
                    abono_val  = float(ab.get("abono", 0) or 0)
                    tag        = f"xls_ab_{idx}"

                    if monto_total > 0:
                        db.add(models.Deuda(
                            tournament_id=tid,
                            player_id=player.id,
                            tipo=tipo_pago if tipo_pago in ("inscripcion","arbitraje") else "manual",
                            fase=fase,
                            monto=monto_total,
                            concepto=concepto,
                            config_tipo=tag,
                        ))

                    if abono_val > 0:
                        db.add(models.Payment(
                            tournament_id=tid,
                            player_id=player.id,
                            payment_type=tipo_pago if tipo_pago in ("inscripcion","arbitraje") else "manual",
                            phase=fase,
                            amount=abono_val,
                            notes=f"{concepto} (Excel)",
                        ))
            else:
                # Formato viejo (retrocompatibilidad)
                CONCEPTOS = [
                    ("inscripcion", "Inscripción torneo",    None),
                    ("arb_f1",      "Arbitrajes Fase 1",     "Fase 1"),
                    ("arb_f2",      "Arbitrajes Fase 2",     "Fase 2"),
                    ("arb_f3",      "Arbitrajes Fase 3",     "Fase 3"),
                ]
                for key, concepto, fase in CONCEPTOS:
                    blk   = j.get(key, {})
                    total = float(blk.get("total", 0) or 0)
                    abono = float(blk.get("abono", 0) or 0)
                    tag   = f"xls_{key}"

                    if total > 0:
                        db.add(models.Deuda(
                            tournament_id=tid, player_id=player.id,
                            tipo="inscripcion" if key == "inscripcion" else "arbitraje",
                            fase=fase, monto=total, concepto=concepto, config_tipo=tag,
                        ))
                    if abono > 0:
                        db.add(models.Payment(
                            tournament_id=tid, player_id=player.id,
                            payment_type="inscripcion" if key == "inscripcion" else "arbitraje",
                            amount=abono, notes=f"{concepto} (Excel)",
                        ))

        except Exception as e:
            errores.append(f"{j.get('nombre','?')}: {str(e)}")

    db.commit()
    return {
        "message": f"✅ {creados} jugadores nuevos · {actualizados} actualizados · {len(errores)} errores",
        "creados": creados, "actualizados": actualizados, "errores": errores,
    }


# ── 5. IMPORTAR JUGADORES (legado) ───────────────────────────────
@router.post("/import-players")
def import_players(body: dict, t: Optional[int] = Query(None), db: Session = Depends(get_db), current_user=Depends(require_admin)):
    jugadores = body.get("jugadores", [])
    if not jugadores:
        raise HTTPException(status_code=400, detail="No hay jugadores para importar")

    if t:
        tid = t
    else:
        active = db.query(models.Tournament).filter(models.Tournament.is_active == True).first()
        if not active:
            raise HTTPException(status_code=400, detail="No hay torneo activo")
        tid = active.id

    creados = 0; actualizados = 0; errores = []

    for j in jugadores:
        try:
            nombre = str(j.get("nombre", "")).strip()
            numero = int(j.get("numero", 0))
            deuda  = float(j.get("deuda_total", 0))
            pagado = float(j.get("pagado", 0))
            if not nombre or not numero:
                continue

            existing = db.query(models.Player).filter(
                models.Player.tournament_id == tid, models.Player.player_number == numero,
            ).first()

            if existing:
                existing.full_name = nombre; player = existing; actualizados += 1
            else:
                player = models.Player(tournament_id=tid, full_name=nombre, player_number=numero, id_number=f"IMPORT-{tid}-{numero}", status="activo")
                db.add(player); db.flush(); creados += 1

            if deuda > 0:
                if not db.query(models.Deuda).filter(models.Deuda.player_id == player.id, models.Deuda.config_tipo == "importado").first():
                    db.add(models.Deuda(tournament_id=tid, player_id=player.id, tipo="manual", monto=deuda, concepto="Deuda importada", config_tipo="importado"))

            if pagado > 0:
                if not db.query(models.Payment).filter(models.Payment.player_id == player.id, models.Payment.notes == "Pago importado").first():
                    db.add(models.Payment(tournament_id=tid, player_id=player.id, payment_type="manual", amount=pagado, notes="Pago importado"))

        except Exception as e:
            errores.append(f"{j.get('nombre','?')}: {str(e)}")

    db.commit()
    return {"message": f"Importación: {creados} creados, {actualizados} actualizados", "creados": creados, "actualizados": actualizados, "errores": errores}


# ── 6. LECTOR DE BOLETÍN DEL TORNEO (PDF) ────────────────────────
@router.post("/read-tournament-pdf")
async def read_tournament_pdf(
    file: UploadFile = File(...),
    current_user=Depends(require_admin),
):
    """
    Lee el PDF del boletín del torneo Metropolitano y extrae
    toda la información relevante para MIRADOR II.
    """
    content  = await file.read()
    filename = (file.filename or "").lower()

    if not filename.endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos PDF")

    try:
        import pypdf
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            reader = pypdf.PdfReader(io.BytesIO(content))
            pages_text = []
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    pages_text.append(t)
            full_text = "\n".join(pages_text)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error leyendo PDF: {str(e)}")

    prompt = f"""Eres un asistente que procesa boletines del Torneo Metropolitano de fútbol amateur colombiano.
Devuelve SOLO un JSON válido sin texto ni backticks.

EXTRAE EXACTAMENTE ESTAS SECCIONES:

{{
  "fase_actual": "Cuartos de Final",
  "fecha_juego": "17 de Mayo de 2026",
  "proxima_fecha": "24 de Mayo de 2026",
  "proximo_partido_mirador": {{
    "rival": "ASTON BIRRA",
    "es_local": true,
    "hora": null,
    "campo": null,
    "fecha_str": "Domingo 24 de Mayo de 2026"
  }},
  "ultimo_resultado_mirador": {{
    "rival": "THE FRIENDS",
    "goles_mirador": 2,
    "goles_rival": 2,
    "es_local": false
  }},
  "resultados": [
    {{"grupo": "A", "local": "THE FRIENDS", "goles_local": 2, "visitante": "MIRADOR II", "goles_visitante": 2, "tiene_mirador": true}},
    {{"grupo": "A", "local": "ASTON BIRRA", "goles_local": 2, "visitante": "DECADENTES F.C.", "goles_visitante": 2, "tiene_mirador": false}},
    {{"grupo": "B", "local": "MANCHESTER F.C.", "goles_local": 2, "visitante": "REAL MAFIA", "goles_visitante": 0, "tiene_mirador": false}},
    {{"grupo": "B", "local": "MACARA F.C.", "goles_local": 2, "visitante": "SATUPSOL F.C.", "goles_visitante": 2, "tiene_mirador": false}}
  ],
  "programacion": [
    {{"grupo": "A", "local": "DECADENTES F.C.", "visitante": "THE FRIENDS", "hora": null, "campo": null, "tiene_mirador": false}},
    {{"grupo": "A", "local": "MIRADOR II", "visitante": "ASTON BIRRA", "hora": null, "campo": null, "tiene_mirador": true}},
    {{"grupo": "B", "local": "SATUPSOL F.C.", "visitante": "MANCHESTER F.C.", "hora": null, "campo": null, "tiene_mirador": false}},
    {{"grupo": "B", "local": "REAL MAFIA", "visitante": "MACARA", "hora": null, "campo": null, "tiene_mirador": false}}
  ],
  "valla_menos_vencida": [
    {{"puesto": 1, "equipo": "MANCHESTER F.C.", "goles_contra": 17, "es_mirador": false}},
    {{"puesto": 2, "equipo": "MACARA", "goles_contra": 18, "es_mirador": false}},
    {{"puesto": 3, "equipo": "MIRADOR II", "goles_contra": 28, "es_mirador": true}}
  ],
  "fair_play": [
    {{"puesto": 1, "equipo": "MIRADOR II", "puntos": 272, "es_mirador": true, "es_ganador": true}},
    {{"puesto": 2, "equipo": "TRIBU BACATA FC", "puntos": 264, "es_mirador": false}}
  ],
  "cronograma": [
    {{"fecha": "24 Mayo", "evento": "Segunda fecha Cuartos de Final"}},
    {{"fecha": "31 Mayo", "evento": "Tercera fecha Cuartos de Final"}},
    {{"fecha": "7 Junio", "evento": "Semifinal"}},
    {{"fecha": "14 Junio", "evento": "Gran Final"}}
  ],
  "tablas_grupo": [
    {{
      "nombre": "GRUPO A",
      "equipos": [
        {{"puesto": 1, "equipo": "MIRADOR II", "pj": 1, "pg": 0, "pe": 1, "pp": 0, "gf": 2, "gc": 2, "puntos": 2, "es_mirador": true}},
        {{"puesto": 2, "equipo": "THE FRIENDS", "pj": 1, "pg": 0, "pe": 1, "pp": 0, "gf": 2, "gc": 2, "puntos": 2, "es_mirador": false}}
      ]
    }},
    {{
      "nombre": "GRUPO B",
      "equipos": [
        {{"puesto": 1, "equipo": "MANCHESTER F.C.", "pj": 1, "pg": 1, "pe": 0, "pp": 0, "gf": 2, "gc": 0, "puntos": 3, "es_mirador": false}}
      ]
    }}
  ],
  "costo_arbitraje": 40000,
  "costo_campo": 140000
}}

REGLAS:
- fase_actual: detecta la fase del torneo (Cuartos de Final, Semifinal, Gran Final)
- resultados: todos los partidos jugados en la última fecha, con goles
- programacion: todos los partidos programados para la próxima fecha (hora y campo null si no aparecen)
- valla_menos_vencida: TODOS los equipos de la lista, ordenados por goles en contra (menos = mejor)
- fair_play: TODOS los equipos de la tabla, con sus puntos
- tablas_grupo: tabla de posiciones de cada grupo con pj, pg, pe, pp, gf, gc, puntos
- tiene_mirador: true si MIRADOR II juega ese partido
- es_mirador: true si el equipo es MIRADOR II
- Si hora o campo no aparecen en el boletín, usa null
- IMPORTANTE: Si el boletín NO tiene programación de próxima fecha para MIRADOR II, pon proximo_partido_mirador: null
- IMPORTANTE: Si el boletín NO tiene resultado de MIRADOR II, pon ultimo_resultado_mirador: null

BOLETÍN:
{full_text[:9000]}
"""


    try:
        c = _client()
        resp = c.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=4000,
            temperature=0.1,
        )
        raw = resp.choices[0].message.content
        raw = raw.strip().replace("```json", "").replace("```", "").strip()
        # Si el JSON viene truncado, intentar cerrarlo
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            # Intentar reparar JSON truncado con varias combinaciones
            repaired = False
            for closing in [']}', ']}]}', '}}', '}}}', ']}}}', '}}]}', ']}}', ']}]}}}']:
                try:
                    data = json.loads(raw + closing)
                    repaired = True
                    break
                except:
                    pass
            if not repaired:
                # Último intento: truncar al último objeto completo
                last_brace = raw.rfind('},')
                if last_brace > 0:
                    try:
                        data = json.loads(raw[:last_brace+1] + ']}}}')
                        repaired = True
                    except:
                        pass
            if not repaired:
                raise HTTPException(status_code=500, detail="Error parseando respuesta de IA: JSON incompleto")
        return {"success": True, "data": data, "texto_extraido": len(full_text)}
    except HTTPException:
        raise
    except Exception as e:
        msg = str(e)
        if "rate_limit_exceeded" in msg or "429" in msg:
            import re
            wait = re.search(r'try again in (\d+m[\d.]+s)', msg)
            wt = f" Espera {wait.group(1)}." if wait else " Intenta en unos minutos."
            raise HTTPException(status_code=429, detail=f"⏳ Límite diario de IA alcanzado.{wt}")
        raise HTTPException(status_code=500, detail=f"Error de IA: {msg}")